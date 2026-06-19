#!/usr/bin/env python3
"""Genera un Excel de horas extra desde un PDF de asistencia.

The script expects attendance PDFs like the ZKTeco-style report used in this
thread: each scan line contains a timestamp and "Registro de entrada" or
"Registro de salida".
"""

from __future__ import annotations

import argparse
import calendar
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

try:
    from pypdf import PdfReader
except ImportError as exc:  # pragma: no cover - user-facing dependency guard
    raise SystemExit(
        "Missing dependency: pypdf. Install it with `python -m pip install pypdf`."
    ) from exc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - user-facing dependency guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `python -m pip install openpyxl`."
    ) from exc


DATETIME_RE = re.compile(
    r"(?P<stamp>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+"
    r"Registro\s+de\s+(?P<kind>entrada|salida)",
    flags=re.IGNORECASE,
)

SPANISH_MONTHS = [
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


@dataclass(frozen=True)
class AttendanceDay:
    day: date
    entry: time | None
    exit: time | None
    status: str
    notes: str


def parse_time(value: str) -> time:
    return datetime.strptime(value, "%H:%M:%S").time()


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_rain_dates(value: str) -> set[date]:
    """Convierte una lista separada por comas en fechas de lluvia."""
    rain_dates: set[date] = set()
    invalid: list[str] = []
    for raw_value in value.split(","):
        item = raw_value.strip()
        if not item:
            continue
        parsed = None
        for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(item, date_format).date()
                break
            except ValueError:
                continue
        if parsed is None:
            invalid.append(item)
        else:
            rain_dates.add(parsed)
    if invalid:
        raise ValueError(
            "Fechas de lluvia inválidas: "
            + ", ".join(invalid)
            + ". Use YYYY-MM-DD o DD/MM/YYYY, separadas por comas."
        )
    return rain_dates


def late_threshold(day: date, late_starts_at: time, rain_dates: set[date]) -> time:
    """Devuelve el inicio de tardanza, sumando 30 minutos en días de lluvia."""
    if day not in rain_dates:
        return late_starts_at
    return (datetime.combine(day, late_starts_at) + timedelta(minutes=30)).time()


def excel_time(value: time | None) -> float | None:
    if value is None:
        return None
    return (value.hour * 3600 + value.minute * 60 + value.second) / 86400


def time_from_hours(base: time, hours: float) -> time:
    base_dt = datetime.combine(date(2000, 1, 1), base)
    return (base_dt + timedelta(hours=hours)).time().replace(microsecond=0)


def monday_of(day: date) -> date:
    return day - timedelta(days=day.weekday())


def decimal_hours(
    start: time | None,
    end: time | None,
    official_entry: time,
) -> float:
    if start is None or end is None:
        return 0.0
    start_minute = start.replace(second=0, microsecond=0)
    end_minute = end.replace(second=0, microsecond=0)
    official_entry_minute = official_entry.replace(second=0, microsecond=0)
    effective_start = max(start_minute, official_entry_minute)
    start_dt = datetime.combine(date(2000, 1, 1), effective_start)
    end_dt = datetime.combine(date(2000, 1, 1), end_minute)
    if end_dt < start_dt:
        end_dt += timedelta(days=1)
    return (end_dt - start_dt).total_seconds() / 3600


def raw_daily_billable(
    start: time | None,
    end: time | None,
    official_entry: time,
    regular_hours: float,
    daily_cap: float,
) -> float:
    return min(daily_cap, max(0.0, decimal_hours(start, end, official_entry) - regular_hours))


def extract_records(pdf_path: Path) -> list[tuple[datetime, str]]:
    reader = PdfReader(str(pdf_path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    records: list[tuple[datetime, str]] = []
    for match in DATETIME_RE.finditer(text):
        stamp = datetime.strptime(match.group("stamp"), "%Y-%m-%d %H:%M:%S")
        kind = match.group("kind").lower()
        records.append((stamp, kind))
    if not records:
        raise ValueError(
            "No attendance scan records were found. Check that the PDF text contains "
            "'Registro de entrada' / 'Registro de salida' rows."
        )
    return sorted(records)


def records_by_day(records: list[tuple[datetime, str]]) -> dict[date, dict[str, list[time]]]:
    grouped: dict[date, dict[str, list[time]]] = defaultdict(
        lambda: {"entrada": [], "salida": []}
    )
    for stamp, kind in records:
        grouped[stamp.date()][kind].append(stamp.time())
    return grouped


def build_actual_days(records: list[tuple[datetime, str]]) -> list[AttendanceDay]:
    grouped = records_by_day(records)
    days: list[AttendanceDay] = []
    for day in sorted(grouped):
        entries = sorted(grouped[day]["entrada"])
        exits = sorted(grouped[day]["salida"])
        entry = entries[0] if entries else None
        exit_time = exits[-1] if exits else None
        if entry and exit_time:
            status = "Real"
            notes = "Registro del PDF"
        elif entry and not exit_time:
            status = "En curso"
            notes = "Entrada encontrada; sin salida en el PDF"
        elif exit_time and not entry:
            status = "Incompleto"
            notes = "Salida encontrada; sin entrada en el PDF"
        else:
            status = "Sin registro"
            notes = ""
        days.append(AttendanceDay(day, entry, exit_time, status, notes))
    return days


def compute_billable_for_known_days(
    days: list[AttendanceDay],
    regular_hours: float,
    daily_cap: float,
    weekly_cap: float,
    monthly_cap: float,
    official_entry: time,
    late_starts_at: time,
    early_exit_before: time,
    allowed_rule_events: int,
    rain_dates: set[date] | None = None,
) -> tuple[dict[date, float], dict[date, int]]:
    rain_dates = rain_dates or set()
    weekly_totals: dict[date, float] = defaultdict(float)
    monthly_total = 0.0
    late_total = 0
    billable: dict[date, float] = {}
    event_counts: dict[date, int] = {}

    for row in sorted(days, key=lambda item: item.day):
        if row.entry is None or row.exit is None:
            billable[row.day] = 0.0
            event_counts[row.day] = 0
            continue

        late_event = int(row.entry >= late_threshold(row.day, late_starts_at, rain_dates))
        early_exit_event = int(row.exit < early_exit_before)
        events = late_event + early_exit_event
        late_total += late_event
        event_counts[row.day] = events

        eligible = late_event == 0 or late_total <= allowed_rule_events
        after_rule = (
            raw_daily_billable(row.entry, row.exit, official_entry, regular_hours, daily_cap)
            if eligible
            else 0.0
        )
        week = monday_of(row.day)
        after_week = min(after_rule, max(0.0, weekly_cap - weekly_totals[week]))
        after_month = min(after_week, max(0.0, monthly_cap - monthly_total))
        weekly_totals[week] += after_week
        monthly_total += after_month
        billable[row.day] = after_month

    return billable, event_counts


def distribute_planned_days(
    actual_days: list[AttendanceDay],
    year: int,
    month: int,
    plan_from: date,
    as_of: date,
    planned_entry: time,
    official_entry: time,
    regular_hours: float,
    daily_cap: float,
    weekly_cap: float,
    monthly_cap: float,
    late_starts_at: time,
    early_exit_before: time,
    allowed_rule_events: int,
    rain_dates: set[date] | None = None,
) -> dict[date, AttendanceDay]:
    known_billable, _ = compute_billable_for_known_days(
        [day for day in actual_days if day.exit is not None],
        regular_hours,
        daily_cap,
        weekly_cap,
        monthly_cap,
        official_entry,
        late_starts_at,
        early_exit_before,
        allowed_rule_events,
        rain_dates,
    )
    actual_total = sum(known_billable.values())
    remaining_month = max(0.0, monthly_cap - actual_total)

    planned: dict[date, AttendanceDay] = {}
    last_day = calendar.monthrange(year, month)[1]
    plan_start_day = max(plan_from, as_of)
    if plan_start_day.year != year or plan_start_day.month != month:
        return planned
    candidates = [
        date(year, month, day)
        for day in range(plan_start_day.day, last_day + 1)
        if date(year, month, day).weekday() < 5
    ]
    if not candidates or remaining_month <= 0:
        return planned

    by_week: dict[date, list[date]] = defaultdict(list)
    for day in candidates:
        by_week[monday_of(day)].append(day)

    weekly_used: dict[date, float] = defaultdict(float)
    for day, hours in known_billable.items():
        weekly_used[monday_of(day)] += hours

    remaining = remaining_month
    for week in sorted(by_week):
        week_room = max(0.0, weekly_cap - weekly_used[week])
        week_target = min(remaining, week_room)
        week_days = by_week[week]
        if week_target <= 0:
            continue
        per_day = week_target / len(week_days)
        for idx, day in enumerate(week_days):
            target = per_day
            if idx == len(week_days) - 1:
                target = week_target - per_day * (len(week_days) - 1)
            target = min(daily_cap, max(0.0, target))
            effective_planned_entry = max(planned_entry, official_entry)
            out_time = time_from_hours(effective_planned_entry, regular_hours + target)
            note = "Planificado para llegar al tope mensual"
            planned[day] = AttendanceDay(day, planned_entry, out_time, "Planificado", note)
        remaining -= week_target
        if remaining <= 1e-9:
            break

    return planned


def make_month_rows(
    actual_days: list[AttendanceDay],
    year: int,
    month: int,
    as_of: date,
    plan_remaining: bool,
    planned_entry: time,
    official_entry: time,
    regular_hours: float,
    daily_cap: float,
    weekly_cap: float,
    monthly_cap: float,
    late_starts_at: time,
    early_exit_before: time,
    allowed_rule_events: int,
    rain_dates: set[date] | None = None,
) -> list[AttendanceDay]:
    actual_by_date = {row.day: row for row in actual_days}
    incomplete_days = [row.day for row in actual_days if row.exit is None and row.entry]
    last_actual_day = max(row.day for row in actual_days)
    incomplete_plannable = [day for day in incomplete_days if day >= as_of]
    plan_from = (
        min(incomplete_plannable)
        if incomplete_plannable
        else max(last_actual_day + timedelta(days=1), as_of)
    )

    planned = (
        distribute_planned_days(
            actual_days,
            year,
            month,
            plan_from,
            as_of,
            planned_entry,
            official_entry,
            regular_hours,
            daily_cap,
            weekly_cap,
            monthly_cap,
            late_starts_at,
            early_exit_before,
            allowed_rule_events,
            rain_dates,
        )
        if plan_remaining
        else {}
    )

    rows: list[AttendanceDay] = []
    last_day = calendar.monthrange(year, month)[1]
    for day_num in range(1, last_day + 1):
        day = date(year, month, day_num)
        if day in actual_by_date and actual_by_date[day].exit is not None:
            rows.append(actual_by_date[day])
        elif day in planned:
            rows.append(planned[day])
        elif day in actual_by_date:
            rows.append(actual_by_date[day])
        elif day.weekday() >= 5:
            rows.append(AttendanceDay(day, None, None, "Fin de semana", ""))
        elif day < as_of:
            rows.append(AttendanceDay(day, None, None, "Ausente", "Sin entrada ni salida"))
        else:
            rows.append(AttendanceDay(day, None, None, "Sin registro", ""))
    return rows


def write_workbook(
    rows: list[AttendanceDay],
    output_path: Path,
    monthly_cap: float,
    regular_hours: float,
    weekly_cap: float,
    daily_cap: float,
    official_entry: time,
    late_starts_at: time,
    early_exit_before: time,
    allowed_rule_events: int,
    rain_dates: set[date] | None = None,
) -> None:
    rain_dates = rain_dates or set()
    wb = Workbook()
    ws = wb.active
    month_name = SPANISH_MONTHS[rows[0].day.month]
    ws.title = f"Detalle {month_name} {rows[0].day.year}"
    principal = wb.create_sheet("Vista principal", 0)

    title_fill = PatternFill("solid", fgColor="DDEAF7")
    panel_fill = PatternFill("solid", fgColor="F3F6F9")
    good_fill = PatternFill("solid", fgColor="EAF4EA")
    header_fill = PatternFill("solid", fgColor="17324D")
    total_fill = PatternFill("solid", fgColor="FFF7D6")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="17324D", bold=True, size=16)
    border = Border(bottom=Side(style="thin", color="CAD3DD"))

    ws.merge_cells("A1:U1")
    ws["A1"] = f"Horas extra - {month_name} {rows[0].day.year}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    assumptions = [
        ("Tope mensual", monthly_cap / 24),
        ("Horas regulares antes de horas extra", regular_hours / 24),
        ("Tope semanal", weekly_cap / 24),
        ("Tope diario", daily_cap / 24),
        ("Hora oficial de entrada", official_entry),
        ("Llegada tarde desde", late_starts_at),
        ("Salida anticipada antes de", early_exit_before),
        ("Usos de cupo permitidos", allowed_rule_events),
        ("Tolerancia adicional por lluvia", timedelta(minutes=30)),
    ]
    for idx, (label, value) in enumerate(assumptions, start=3):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)
        ws.cell(idx, 1).font = Font(bold=True)
        ws.cell(idx, 1).fill = panel_fill
        ws.cell(idx, 2).fill = panel_fill
    for cell in ("B3", "B4", "B5", "B6"):
        ws[cell].number_format = "[h]:mm:ss"
    ws["B7"].number_format = "hh:mm:ss"
    ws["B8"].number_format = "hh:mm:ss"
    ws["B9"].number_format = "hh:mm:ss"
    ws["B10"].number_format = "0"
    ws["B11"].number_format = "[h]:mm:ss"

    header_row = 13
    first_data_row = header_row + 1
    last_data_row = header_row + len(rows)
    previous_row = header_row

    metrics = [
        ("Total horas extra", f"=SUM(R{first_data_row}:R{last_data_row})"),
        ("Horas extra reales", f'=SUMIFS(R{first_data_row}:R{last_data_row},D{first_data_row}:D{last_data_row},"Real")'),
        ("Horas extra planificadas", f'=SUMIFS(R{first_data_row}:R{last_data_row},D{first_data_row}:D{last_data_row},"Planificado")'),
        ("Restante al tope mensual", "=MAX(0,$B$3-E3)"),
        ("Usos de cupo registrados", f"=SUM(L{first_data_row}:L{last_data_row})"),
        ("Usos de cupo restantes", "=MAX(0,$B$10-E7)"),
        ("Llegadas tarde registradas", f'=COUNTIF(J{first_data_row}:J{last_data_row},"Sí")'),
    ]
    for idx, (label, formula) in enumerate(metrics, start=3):
        ws.cell(idx, 4, label)
        ws.cell(idx, 5, formula)
        ws.cell(idx, 4).font = Font(bold=True)
        ws.cell(idx, 4).fill = good_fill
        ws.cell(idx, 5).fill = good_fill
        ws.cell(idx, 5).number_format = "[h]:mm:ss" if idx <= 6 else "0"

    headers = [
        "Fecha",
        "Inicio de semana",
        "Día",
        "Estado",
        "Hora de entrada",
        "Hora de salida",
        "Horas trabajadas",
        "Horas extra brutas",
        "¿Día de lluvia?",
        "¿Llegada tarde?",
        "¿Salida anticipada?",
        "Uso de cupo",
        "Cupo acumulado",
        "Llegadas tarde acumuladas",
        "¿Elegible para horas extra?",
        "Después de regla",
        "Después de tope semanal",
        "Horas extra mensual",
        "Acumulado mensual",
        "Tope restante",
        "Notas",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    day_names = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
    for offset, row in enumerate(rows, start=first_data_row):
        ws.cell(offset, 1, row.day)
        ws.cell(offset, 2, monday_of(row.day))
        ws.cell(offset, 3, day_names[row.day.weekday()])
        ws.cell(offset, 4, row.status)
        ws.cell(offset, 5, row.entry)
        ws.cell(offset, 6, row.exit)
        ws.cell(offset, 7, f'=IF(OR(E{offset}="",F{offset}=""),0,MOD(FLOOR(F{offset}*1440,1)-MAX(FLOOR(E{offset}*1440,1),FLOOR($B$7*1440,1)),1440)/1440)')
        ws.cell(offset, 8, f"=MIN($B$6,MAX(0,G{offset}-$B$4))")
        ws.cell(offset, 9, "Sí" if row.day in rain_dates else "No")
        ws.cell(offset, 10, f'=IF(AND(E{offset}<>"",E{offset}>=$B$8+IF(I{offset}="Sí",$B$11,0)),"Sí","No")')
        ws.cell(offset, 11, f'=IF(AND(F{offset}<>"",F{offset}<$B$9),"Sí","No")')
        ws.cell(offset, 12, f'=IF(J{offset}="Sí",1,0)+IF(K{offset}="Sí",1,0)')
        ws.cell(offset, 13, f"=SUM($L${first_data_row}:L{offset})")
        ws.cell(offset, 14, f'=COUNTIF($J${first_data_row}:J{offset},"Sí")')
        ws.cell(offset, 15, f'=IF(OR(J{offset}<>"Sí",N{offset}<=$B$10),"Sí","No")')
        ws.cell(offset, 16, f'=IF(O{offset}="Sí",H{offset},0)')
        ws.cell(
            offset,
            17,
            f"=MIN(P{offset},MAX(0,$B$5-SUMIFS($Q${previous_row}:Q{offset - 1},$B${previous_row}:B{offset - 1},B{offset})))",
        )
        ws.cell(offset, 18, f"=MIN(Q{offset},MAX(0,$B$3-SUM($R${previous_row}:R{offset - 1})))")
        ws.cell(offset, 19, f"=SUM($R${first_data_row}:R{offset})")
        ws.cell(offset, 20, f"=MAX(0,$B$3-S{offset})")
        ws.cell(offset, 21, row.notes)

    for row in ws.iter_rows(min_row=header_row, max_row=last_data_row, min_col=1, max_col=21):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in (1, 2):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=first_data_row, max_row=last_data_row):
            for item in cell:
                item.number_format = "mmm d, yyyy"
    for col in (5, 6):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=first_data_row, max_row=last_data_row):
            for item in cell:
                item.number_format = "hh:mm:ss"
    for col in (7, 8, 16, 17, 18, 19, 20):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=first_data_row, max_row=last_data_row):
            for item in cell:
                item.number_format = "[h]:mm:ss"
    for col in (12, 13, 14):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=first_data_row, max_row=last_data_row):
            for item in cell:
                item.number_format = "0"
    for cell in ws[last_data_row]:
        cell.fill = total_fill

    widths = {
        "A": 14,
        "B": 14,
        "C": 8,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 13,
        "J": 13,
        "K": 13,
        "L": 12,
        "M": 14,
        "N": 15,
        "O": 12,
        "P": 14,
        "Q": 14,
        "R": 14,
        "S": 13,
        "T": 13,
        "U": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{first_data_row}"

    principal.merge_cells("A1:N1")
    principal["A1"] = f"Horas extra - {month_name} {rows[0].day.year}"
    principal["A1"].font = title_font
    principal["A1"].fill = title_fill
    principal["A1"].alignment = Alignment(horizontal="center")

    for idx in range(3, 12):
        principal.cell(idx, 1, ws.cell(idx, 1).value)
        principal.cell(idx, 2, ws.cell(idx, 2).value)
        principal.cell(idx, 1).font = Font(bold=True)
        principal.cell(idx, 1).fill = panel_fill
        principal.cell(idx, 2).fill = panel_fill
        principal.cell(idx, 2).number_format = ws.cell(idx, 2).number_format
    for idx in range(3, 10):
        principal.cell(idx, 4, ws.cell(idx, 4).value)
        principal.cell(idx, 5, f"='{ws.title}'!E{idx}")
        principal.cell(idx, 4).font = Font(bold=True)
        principal.cell(idx, 4).fill = good_fill
        principal.cell(idx, 5).fill = good_fill
        principal.cell(idx, 5).number_format = ws.cell(idx, 5).number_format

    principal.merge_cells("H2:N2")
    principal["H2"] = "Resumen semanal de horas extra"
    principal["H2"].font = Font(color="17324D", bold=True, size=12)
    principal["H2"].fill = title_fill
    principal["H2"].alignment = Alignment(horizontal="center")
    summary_headers = [
        "Inicio de semana",
        "Horas extra reales",
        "Horas extra planificadas",
        "Total horas extra",
        "Tope semanal",
        "Tope semanal sin usar",
        "Notas",
    ]
    for col, header in enumerate(summary_headers, start=8):
        cell = principal.cell(3, col, header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    weeks = sorted({monday_of(row.day) for row in rows})
    for idx, week in enumerate(weeks, start=4):
        principal.cell(idx, 8, week)
        principal.cell(idx, 9, f'=SUMIFS(\'{ws.title}\'!$R${first_data_row}:$R${last_data_row},\'{ws.title}\'!$B${first_data_row}:$B${last_data_row},H{idx},\'{ws.title}\'!$D${first_data_row}:$D${last_data_row},"Real")')
        principal.cell(idx, 10, f'=SUMIFS(\'{ws.title}\'!$R${first_data_row}:$R${last_data_row},\'{ws.title}\'!$B${first_data_row}:$B${last_data_row},H{idx},\'{ws.title}\'!$D${first_data_row}:$D${last_data_row},"Planificado")')
        principal.cell(idx, 11, f"=SUM(I{idx}:J{idx})")
        principal.cell(idx, 12, f"='{ws.title}'!$B$5")
        principal.cell(idx, 13, f"=MAX(0,L{idx}-K{idx})")
        principal.cell(idx, 14, f'=IF(K{idx}=0,"Sin registros",IF(M{idx}=0,"En tope semanal","Debajo del tope semanal"))')
    for row in principal.iter_rows(min_row=3, max_row=3 + len(weeks), min_col=8, max_col=14):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in principal.iter_cols(min_col=8, max_col=8, min_row=4, max_row=3 + len(weeks)):
        for item in cell:
            item.number_format = "mmm d, yyyy"
    for cell in principal.iter_cols(min_col=9, max_col=13, min_row=4, max_row=3 + len(weeks)):
        for item in cell:
            item.number_format = "[h]:mm:ss"

    principal_source_columns = (1, 2, 3, 4, 5, 6, 7, 8, 19, 21)
    for target_col, source_col in enumerate(principal_source_columns, start=1):
        header = principal.cell(header_row, target_col, ws.cell(header_row, source_col).value)
        header.fill = header_fill
        header.font = white_font
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        principal.column_dimensions[get_column_letter(target_col)].width = ws.column_dimensions[get_column_letter(source_col)].width
        for row_idx in range(first_data_row, last_data_row + 1):
            source_ref = f"'{ws.title}'!{get_column_letter(source_col)}{row_idx}"
            principal.cell(row_idx, target_col, f'=IF({source_ref}="","",{source_ref})')
            principal.cell(row_idx, target_col).number_format = ws.cell(row_idx, source_col).number_format
            principal.cell(row_idx, target_col).border = border
            principal.cell(row_idx, target_col).alignment = Alignment(vertical="center", wrap_text=True)
    for cell in principal[last_data_row][:10]:
        cell.fill = total_fill
    for col in range(8, 15):
        principal.column_dimensions[get_column_letter(col)].width = max(
            principal.column_dimensions[get_column_letter(col)].width or 0, 18
        )
    principal.column_dimensions["N"].width = 22
    principal.freeze_panes = f"A{first_data_row}"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crea un Excel de horas extra desde un PDF de asistencia."
    )
    parser.add_argument("pdf", type=Path, help="PDF de asistencia de entrada")
    parser.add_argument("-o", "--output", type=Path, help="Ruta del archivo .xlsx de salida")
    parser.add_argument("--year", type=int, help="Año del reporte. Por defecto usa el primer registro del PDF.")
    parser.add_argument("--month", type=int, help="Mes del reporte. Por defecto usa el primer registro del PDF.")
    parser.add_argument("--monthly-cap", type=float, default=25.0)
    parser.add_argument("--weekly-cap", type=float, default=8.0)
    parser.add_argument("--daily-cap", type=float, default=3.0)
    parser.add_argument("--regular-hours", type=float, default=8.0)
    parser.add_argument("--official-entry", default="08:00:00")
    parser.add_argument("--late-starts-at", default="08:16:00")
    parser.add_argument("--early-exit-before", default="16:00:00")
    parser.add_argument("--rule-events-allowed", type=int, default=3)
    parser.add_argument("--planned-entry", default="08:10:04")
    parser.add_argument(
        "--rain-dates",
        type=parse_rain_dates,
        default=set(),
        help="Fechas de lluvia separadas por comas (YYYY-MM-DD o DD/MM/YYYY).",
    )
    parser.add_argument(
        "--as-of",
        type=parse_date,
        default=date.today(),
        help="Fecha desde la cual se permite planificar, en formato YYYY-MM-DD. Por defecto usa hoy.",
    )
    parser.add_argument(
        "--no-plan",
        action="store_true",
        help="No agregar días planificados para llegar al tope mensual.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    records = extract_records(args.pdf)
    year = args.year or records[0][0].year
    month = args.month or records[0][0].month
    actual_days = build_actual_days(records)
    rows = make_month_rows(
        actual_days=actual_days,
        year=year,
        month=month,
        as_of=args.as_of,
        plan_remaining=not args.no_plan,
        planned_entry=parse_time(args.planned_entry),
        official_entry=parse_time(args.official_entry),
        regular_hours=args.regular_hours,
        daily_cap=args.daily_cap,
        weekly_cap=args.weekly_cap,
        monthly_cap=args.monthly_cap,
        late_starts_at=parse_time(args.late_starts_at),
        early_exit_before=parse_time(args.early_exit_before),
        allowed_rule_events=args.rule_events_allowed,
        rain_dates=args.rain_dates,
    )
    output = args.output
    if output is None:
        output = args.pdf.with_name(f"{args.pdf.stem}_horas_extra.xlsx")
    write_workbook(
        rows=rows,
        output_path=output,
        monthly_cap=args.monthly_cap,
        regular_hours=args.regular_hours,
        weekly_cap=args.weekly_cap,
        daily_cap=args.daily_cap,
        official_entry=parse_time(args.official_entry),
        late_starts_at=parse_time(args.late_starts_at),
        early_exit_before=parse_time(args.early_exit_before),
        allowed_rule_events=args.rule_events_allowed,
        rain_dates=args.rain_dates,
    )
    print(output)


if __name__ == "__main__":
    main()
