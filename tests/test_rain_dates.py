import sys
import tempfile
import unittest
from datetime import date, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from generate_billable_hours_from_pdf import (  # noqa: E402
    AttendanceDay,
    compute_billable_for_known_days,
    parse_rain_dates,
    write_workbook,
)


class RainDatesTests(unittest.TestCase):
    def test_parses_both_supported_formats(self):
        self.assertEqual(
            parse_rain_dates("2026-05-04, 12/05/2026:45"),
            {date(2026, 5, 4): 30, date(2026, 5, 12): 45},
        )

    def test_rejects_invalid_date_with_clear_message(self):
        with self.assertRaisesRegex(ValueError, "Fechas o tolerancias de lluvia inválidas: 31/02/2026"):
            parse_rain_dates("31/02/2026")

    def test_rejects_tolerance_below_standard_minimum(self):
        with self.assertRaisesRegex(ValueError, "tolerancia mínima es 30 minutos"):
            parse_rain_dates("04/05/2026:29")

    def test_normal_and_rain_thresholds_preserve_billable_hours(self):
        days = [
            AttendanceDay(date(2026, 5, 4), time(8, 30), time(19, 0), "Real", ""),
            AttendanceDay(date(2026, 5, 5), time(8, 45), time(19, 0), "Real", ""),
            AttendanceDay(date(2026, 5, 6), time(8, 46), time(19, 0), "Real", ""),
        ]
        billable, events = compute_billable_for_known_days(
            days, 8, 3, 8, 25, time(8, 0), time(8, 16), time(16, 0), 3,
            {date(2026, 5, 5): 30, date(2026, 5, 6): 30},
        )
        self.assertEqual(events[date(2026, 5, 4)], 1)
        self.assertEqual(events[date(2026, 5, 5)], 0)
        self.assertEqual(events[date(2026, 5, 6)], 1)
        self.assertGreater(billable[date(2026, 5, 5)], 0)

    def test_each_rain_day_uses_its_own_extended_tolerance(self):
        days = [
            AttendanceDay(date(2026, 5, 11), time(9, 0), time(19, 0), "Real", ""),
            AttendanceDay(date(2026, 5, 12), time(9, 1), time(19, 0), "Real", ""),
        ]
        _, events = compute_billable_for_known_days(
            days, 8, 3, 8, 25, time(8, 0), time(8, 16), time(16, 0), 3,
            {date(2026, 5, 11): 45, date(2026, 5, 12): 45},
        )
        self.assertEqual(events[date(2026, 5, 11)], 0)
        self.assertEqual(events[date(2026, 5, 12)], 1)

    def test_rain_arrival_does_not_consume_quota_or_trigger_fourth_late_exclusion(self):
        days = [
            AttendanceDay(date(2026, 5, day), time(8, 20), time(19, 0), "Real", "")
            for day in (4, 5, 6)
        ]
        days.extend(
            [
                AttendanceDay(date(2026, 5, 7), time(8, 45), time(19, 0), "Real", ""),
                AttendanceDay(date(2026, 5, 8), time(8, 20), time(19, 0), "Real", ""),
            ]
        )
        billable, events = compute_billable_for_known_days(
            days, 8, 3, 20, 25, time(8, 0), time(8, 16), time(16, 0), 3,
            {date(2026, 5, 7): 30},
        )
        self.assertEqual(events[date(2026, 5, 7)], 0)
        self.assertGreater(billable[date(2026, 5, 7)], 0)
        self.assertEqual(billable[date(2026, 5, 8)], 0)

    def test_workbook_opens_with_compact_view_and_keeps_full_detail(self):
        from openpyxl import load_workbook

        rows = [
            AttendanceDay(date(2026, 5, 4), time(8, 10), time(19, 0), "Real", "Registro")
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "horas_extra.xlsx"
            write_workbook(
                rows, output, 25, 8, 8, 3, time(8, 0), time(8, 16),
                time(16, 0), 3, {},
            )
            workbook = load_workbook(output, data_only=False)

        self.assertEqual(workbook.sheetnames, ["Vista principal", "Detalle Mayo 2026"])
        principal = workbook["Vista principal"]
        detail = workbook["Detalle Mayo 2026"]
        source_columns = (1, 2, 3, 4, 5, 6, 7, 8, 19, 21)
        self.assertEqual(
            [principal.cell(13, col).value for col in range(1, 11)],
            [detail.cell(13, col).value for col in source_columns],
        )
        self.assertEqual(
            [principal.cell(3, col).value for col in range(8, 15)],
            [
                "Inicio de semana", "Horas extra reales", "Horas extra planificadas",
                "Total horas extra", "Tope semanal", "Tope semanal sin usar", "Notas",
            ],
        )
        self.assertIn("'Detalle Mayo 2026'!S14", principal["I14"].value)
        self.assertEqual(detail["V13"].value, "Tolerancia adicional por lluvia")


if __name__ == "__main__":
    unittest.main()
